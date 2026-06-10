from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import database


def _score_to_point(score):
    score = float(score)
    if score >= 90:
        return 4.0
    if score >= 80:
        return 3.0
    if score >= 70:
        return 2.0
    if score >= 60:
        return 1.0
    return 0.0


def _build_report_rows():
    students = database.get_students()
    courses = database.get_courses()
    scores = database.fetch_all(
        """
        SELECT sc.student_id, sc.course_id, sc.score, c.credit
        FROM scores sc
        JOIN courses c ON c.course_id = sc.course_id
        """
    )
    grouped = {}
    for row in scores:
        grouped.setdefault(row["student_id"], []).append(row)

    rows = []
    for student in students:
        student_scores = grouped.get(student["student_id"], [])
        score_map = {row["course_id"]: row["score"] for row in student_scores}
        if student_scores:
            avg_score = sum(row["score"] for row in student_scores) / len(student_scores)
            total_credit = sum(row["credit"] for row in student_scores)
            gpa = (
                sum(_score_to_point(row["score"]) * row["credit"] for row in student_scores)
                / total_credit
                if total_credit
                else 0
            )
            warning = "需预警" if avg_score < 60 else "正常"
        else:
            avg_score = None
            gpa = None
            warning = "暂无成绩"

        rows.append(
            {
                "student_id": student["student_id"],
                "name": student["name"],
                "class_name": student["class_name"],
                "scores": score_map,
                "avg_score": avg_score,
                "gpa": gpa,
                "warning": warning,
            }
        )
    return courses, rows


def export_grade_report(path):
    courses, rows = _build_report_rows()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "学生成绩报表"

    headers = ["学号", "姓名", "班级"]
    headers.extend([row["course_name"] for row in courses])
    headers.extend(["平均分", "GPA", "是否预警"])
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    warning_fill = PatternFill("solid", fgColor="F9D6D5")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_data in rows:
        row = [
            row_data["student_id"],
            row_data["name"],
            row_data["class_name"],
        ]
        for course in courses:
            score = row_data["scores"].get(course["course_id"])
            row.append("" if score is None else score)
        row.append("" if row_data["avg_score"] is None else round(row_data["avg_score"], 2))
        row.append("" if row_data["gpa"] is None else round(row_data["gpa"], 2))
        row.append(row_data["warning"])
        sheet.append(row)
        if row_data["warning"] == "需预警":
            for cell in sheet[sheet.max_row]:
                cell.fill = warning_fill

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        column_letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[column_letter].width = max(12, max_length + 2)

    workbook.save(path)
